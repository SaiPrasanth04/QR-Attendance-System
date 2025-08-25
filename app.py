import os
import pandas as pd
import json
from flask import Flask, render_template, request, redirect, url_for, send_file, send_from_directory, flash, jsonify, session
from utils.excel_handler import (
    ensure_workbook, get_batches, read_batch_df, add_student, mark_present, mark_absent_all, download_path, ensure_email_column_all_batches
)
from utils.qr_handler import (
    generate_qr_for_student, generate_qr_for_batch, zip_qr_for_batch
)
from apscheduler.schedulers.background import BackgroundScheduler
from utils.email_handler import send_absent_email
# Email config (set your real credentials or use environment variables)
SENDER_EMAIL = "saiprasanth.d.acet@gmail.com"
SENDER_PASSWORD = "huxdmfdieygjduei"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

app = Flask(__name__)
app.secret_key = "super-secret-key"  # replace in production


@app.route('/edit-batch', methods=['POST'])
def edit_batch():
    from utils.user_handler import authenticate_user, set_user_batch
    import openpyxl
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'mentor') if user else 'mentor'
    if role == 'admin':
        excel_file = EXCEL_FILE
    else:
        excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
    old_batch = request.form.get('old_batch_name', '').strip()
    new_batch = request.form.get('new_batch_name', '').strip()
    if not old_batch or not new_batch or old_batch == new_batch:
        flash('Invalid batch name.', 'danger')
        return redirect(url_for('index'))
    try:
        wb = openpyxl.load_workbook(excel_file)
        if old_batch not in wb.sheetnames:
            flash(f'Batch {old_batch} not found.', 'danger')
            return redirect(url_for('index'))
        if new_batch in wb.sheetnames:
            flash(f'Batch {new_batch} already exists.', 'danger')
            return redirect(url_for('index'))
        ws = wb[old_batch]
        ws.title = new_batch
        try:
            wb.save(excel_file)
        except PermissionError:
            flash('Cannot save changes. Please close the Excel file in all other programs and try again.', 'danger')
            return redirect(url_for('index'))
        # Update batch list for mentor
        if role != 'admin':
            batch_list = [b.strip() for b in user.get('batch', '').split(',') if b.strip()] if user else []
            batch_list = [new_batch if b == old_batch else b for b in batch_list]
            set_user_batch(username, ','.join(batch_list))
        flash(f'Batch renamed from {old_batch} to {new_batch}.', 'success')
    except PermissionError:
        flash('Cannot save changes. Please close the Excel file in all other programs and try again.', 'danger')
    except Exception as e:
        flash(f'Error renaming batch: {e}', 'danger')
    return redirect(url_for('index'))

# Route to send mail to today's absentees only
@app.route("/send-absent-mail-today", methods=["POST"])
def send_absent_mail_today():
    from utils.user_handler import authenticate_user
    from utils.excel_handler import read_batch_df
    from utils.email_handler import send_absent_email
    from datetime import date
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    if role == 'admin':
        excel_file = EXCEL_FILE
        batch_list = get_batches(excel_file)
    else:
        excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
        batch_list = [b.strip() for b in user.get('batch', '').split(',') if b.strip()] if user else []
    today = date.today().strftime("%Y-%m-%d")
    sent_count = 0
    for batch in batch_list:
        try:
            df = read_batch_df(excel_file, batch)
            if today in df.columns:
                for _, row in df.iterrows():
                    if str(row.get(today, '')).strip().lower() == 'absent':
                        student_name = row.get('Name', 'Student')
                        student_email = row.get('Email', '')
                        if student_email:
                            if send_absent_email(
                                to_email=student_email,
                                student_name=student_name,
                                date_str=today,
                                batch=batch,
                                sender_email=SENDER_EMAIL,
                                sender_password=SENDER_PASSWORD,
                                smtp_server=SMTP_SERVER,
                                smtp_port=SMTP_PORT
                            ):
                                sent_count += 1
        except Exception:
            pass
    flash(f"Sent absent emails to {sent_count} students marked absent today.", "success")
    return redirect(url_for("index"))


# Route to delete a student by ID and batch
@app.route("/delete-student", methods=["POST"])
def delete_student_route():
    from utils.user_handler import authenticate_user
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    if role == 'admin':
        excel_file = EXCEL_FILE
    else:
        excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
    student_id = request.form.get("delete_student_id", "").strip()
    batch_name = request.form.get("delete_batch_name", "").strip()
    if not (student_id and batch_name):
        flash("Please provide Student ID and Batch.", "danger")
        return redirect(url_for("index"))
    try:
        from utils.excel_handler import delete_student
        deleted = delete_student(excel_file, batch_name, student_id)
        if deleted:
            flash(f"Student {student_id} removed from {batch_name}.", "success")
        else:
            flash(f"Student {student_id} not found in {batch_name}.", "warning")
    except Exception as e:
        flash(f"Error deleting student: {e}", "danger")
    return redirect(url_for("index"))



@app.route('/upload-batch', methods=['GET', 'POST'])
def upload_batch():
    from utils.user_handler import authenticate_user, set_user_batch
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    if role == 'admin':
        flash('Admins should upload using the dashboard.', 'warning')
        return redirect(url_for('index'))
    excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
    if request.method == 'POST' and 'batch_file' in request.files:
        f = request.files['batch_file']
        if f and f.filename.lower().endswith('.xlsx'):
            f.save(excel_file)
            # Add all sheet names to teacher's batch list
            xl = pd.ExcelFile(excel_file, engine="openpyxl")
            for sheet in xl.sheet_names:
                set_user_batch(username, sheet)
            flash('Batch Excel uploaded and batches assigned!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Please upload a .xlsx file.', 'danger')
    return render_template('upload_batch.html')

# Logout route
@app.route('/logout')
def logout():
    session.pop('user', None)
    flash('Logged out.', 'info')
    return redirect(url_for('login'))


# --- Auth helpers and routes ---
from utils.user_handler import add_user, authenticate_user

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = authenticate_user(email, password, by_email=True)
        if user:
            session['user'] = user['username']
            flash('Login successful!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid email or password.', 'danger')
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        role = request.form.get('role', 'teacher')
        if add_user(username, email, password, role=role):
            flash('Account created! Please log in.', 'success')
            return redirect(url_for('login'))
        else:
            flash('Username or email already exists.', 'danger')
    return render_template('signup.html')



# --- Auth helpers and routes ---
@app.route('/login', methods=['GET', 'POST'])


@app.route('/signup', methods=['GET', 'POST'])

@app.route('/logout')
## (removed all login/signup/logout routes)

@app.route("/download/batch/<batch>/qr-pdf")
def download_qr_pdf(batch):
    try:
        from utils.user_handler import authenticate_user
        username = session.get('user')
        user = authenticate_user(username, None)
        role = user.get('role', 'teacher') if user else 'teacher'
        if role == 'admin':
            excel_file = EXCEL_FILE
        else:
            excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
        from utils.qr_handler import pdf_qr_for_batch
        pdf_path = pdf_qr_for_batch(excel_file, batch, QR_DIR, DATA_DIR)
        return send_file(pdf_path, as_attachment=True)
    except Exception as e:
        flash(f"Error creating PDF: {e}", "danger")
        return redirect(url_for("view_batch_qr", batch=batch))

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
STATIC_DIR = os.path.join(BASE_DIR, "static")
QR_DIR = os.path.join(STATIC_DIR, "qr_codes")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(QR_DIR, exist_ok=True)

EXCEL_FILE = os.path.join(BASE_DIR, "data", "attendance.xlsx")

from utils.excel_handler import ensure_email_column_all_batches
# Ensure workbook exists
ensure_workbook(EXCEL_FILE)
ensure_email_column_all_batches(EXCEL_FILE)



@app.route("/", methods=["GET", "POST"])
def index():
    if 'user' not in session:
        return redirect(url_for('login'))
    import shutil
    from utils.user_handler import authenticate_user, set_user_batch
    username = session['user']
    # Get user info
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    batch = user.get('batch', '') if user else ''
    if not isinstance(batch, str):
        batch = ''
    # Excel file path logic
    if role == 'admin':
        excel_file = EXCEL_FILE
    else:
        excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
        if not os.path.exists(excel_file):
            from utils.excel_handler import ensure_workbook
            ensure_workbook(excel_file)
    # Upload/replace Excel (admin only)
    if request.method == "POST" and "excel_file" in request.files and role == 'admin':
        f = request.files["excel_file"]
        if f and f.filename.lower().endswith(".xlsx"):
            f.save(excel_file)
            for batch_folder in os.listdir(QR_DIR):
                batch_path = os.path.join(QR_DIR, batch_folder)
                if os.path.isdir(batch_path):
                    shutil.rmtree(batch_path)
            flash("Excel file uploaded successfully. Old QR codes removed.", "success")
            return redirect(url_for("index"))
        else:
            flash("Please upload a .xlsx file.", "danger")
            return redirect(url_for("index"))
    # Get batches
    all_batches = []
    try:
        all_batches = get_batches(excel_file)
    except Exception as e:
        flash(f"Error reading Excel file: {e}. Please upload a valid .xlsx file.", "danger")
    # For teachers, show only batches in their batch list; for admin, show all
    if role == 'admin':
        batches = all_batches
    else:
        batch_list = [b.strip() for b in batch.split(',') if b.strip()] if batch else []
        batches = [b for b in all_batches if b in batch_list]
    # Build batch_details for template
    batch_details = []
    for b in batches:
        try:
            df = read_batch_df(excel_file, b)
            students = df.to_dict(orient='records')
            batch_details.append({'batch': b, 'students': students})
        except Exception:
            batch_details.append({'batch': b, 'students': []})
    return render_template("index.html", batches=batches, batch_details=batch_details, role=role)



@app.route("/add-student", methods=["POST"])
def add_student_route():
    from utils.user_handler import authenticate_user, set_user_batch
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    batch = user.get('batch', '') if user else ''
    # Use correct Excel file
    if role == 'admin':
        excel_file = EXCEL_FILE
    else:
        excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
        if not os.path.exists(excel_file):
            from utils.excel_handler import ensure_workbook
            ensure_workbook(excel_file)

    student_id = request.form.get("student_id", "").strip()
    student_name = request.form.get("student_name", "").strip()
    student_email = request.form.get("email", "").strip()
    batch_name = request.form.get("batch_name", "").strip()
    gen_qr = request.form.get("gen_qr") == "on"

    if not (student_id and student_name and batch_name and student_email):
        flash("Please provide ID, Name, Email, and Batch.", "danger")
        return redirect(url_for("index"))

    try:
        added = add_student(excel_file, batch_name, student_id, student_name, student_email)
        if not added:
            # Check if duplicate email exists in any batch
            xl = pd.ExcelFile(excel_file, engine="openpyxl")
            duplicate_email = False
            for sheet in xl.sheet_names:
                sdf = pd.read_excel(excel_file, sheet_name=sheet, engine="openpyxl")
                if "Email" in sdf.columns and (sdf["Email"].astype(str).str.lower() == student_email.lower()).any():
                    duplicate_email = True
                    break
            if duplicate_email:
                flash(f"A student with the email {student_email} already exists in another batch. Each student must have a unique email.", "danger")
            else:
                flash(f"Student with ID {student_id} already exists in {batch_name}.", "warning")
        else:
            flash(f"Student {student_name} ({student_id}) added to {batch_name}.", "success")
            if gen_qr:
                path = generate_qr_for_student(student_id, student_name, batch_name, QR_DIR)
                flash(f"QR generated: {path}", "success")
            # Always add batch to teacher's batch list if not already present
            if role == 'teacher':
                set_user_batch(username, batch_name)
    except Exception as e:
        flash(f"Error adding student: {e}", "danger")
    return redirect(url_for("index"))


@app.route("/generate-qr/<batch>")
def generate_qr_batch(batch):
    from utils.user_handler import authenticate_user
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    batch_list = [b.strip() for b in user.get('batch', '').split(',') if b.strip()] if user else []
    if role == 'admin':
        excel_file = EXCEL_FILE
    else:
        excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
    # Only allow teacher to generate QR for batches in their batch list
    if role != 'admin' and batch not in batch_list:
        flash("Access denied: You can only generate QR for your own batches.", "danger")
        return redirect(url_for("index"))
    try:
        out_dir, count = generate_qr_for_batch(excel_file, batch, QR_DIR)
        flash(f"Generated {count} QR codes for batch {batch}.", "success")
        return redirect(url_for("view_batch_qr", batch=batch))
    except Exception as e:
        flash(f"Error generating QR: {e}", "danger")
        return redirect(url_for("index"))


@app.route("/batch/<batch>/qrs")
def view_batch_qr(batch):
    folder = os.path.join(QR_DIR, batch)
    files = []
    if os.path.isdir(folder):
        files = [f for f in os.listdir(folder) if f.lower().endswith(".png")]
        files.sort()
    return render_template("batch_qr.html", batch=batch, files=files)


@app.route("/download/batch/<batch>/qr-zip")
def download_qr_zip(batch):
    from utils.user_handler import authenticate_user
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    batch_list = [b.strip() for b in user.get('batch', '').split(',') if b.strip()] if user else []
    if role != 'admin' and batch not in batch_list:
        flash("Access denied: You can only download QR ZIP for your own batch.", "danger")
        return redirect(url_for("index"))
    try:
        zip_path = zip_qr_for_batch(batch, QR_DIR, DATA_DIR)
        return send_file(zip_path, as_attachment=True)
    except Exception as e:
        flash(f"Error creating ZIP: {e}", "danger")
        return redirect(url_for("view_batch_qr", batch=batch))


@app.route("/attendance/<batch>")
def attendance_batch(batch):
    from utils.user_handler import authenticate_user
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    batch_list = [b.strip() for b in user.get('batch', '').split(',') if b.strip()] if user else []
    if role == 'admin':
        excel_file = EXCEL_FILE
    else:
        excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
    # Only allow teacher to access batches in their batch list
    if role != 'admin' and batch not in batch_list:
        flash("Access denied: You can only view your own batches.", "danger")
        return redirect(url_for("index"))
    try:
        import calendar
        from datetime import datetime
        df = read_batch_df(excel_file, batch)
        # Identify date columns (assume date columns are after Email)
        date_cols = []
        for col in df.columns:
            try:
                datetime.strptime(str(col), "%Y-%m-%d")
                date_cols.append(col)
            except Exception:
                pass
        # Filter date columns for current month
        now = datetime.now()
        current_month = now.strftime("%Y-%m")
        month_date_cols = [col for col in date_cols if str(col).startswith(current_month)]
        # Calculate present count for each student
        present_counts = []
        for idx, row in df.iterrows():
            present = 0
            for col in month_date_cols:
                val = str(row.get(col, "")).strip().lower()
                if val == "present":
                    present += 1
            present_counts.append(present)
        total_days = len(month_date_cols)
        # Add summary columns to display
        headers = list(df.columns) + ["Present (This Month)", "Total Days (This Month)"]
        rows = [list(row) + [present_counts[i], total_days] for i, row in enumerate(df.values.tolist())]
        return render_template("attendance.html", batch=batch, headers=headers, rows=rows)
    except Exception as e:
        flash(f"Error reading batch: {e}", "danger")
        return redirect(url_for("index"))


@app.route("/download/excel")
def download_excel():
    from utils.user_handler import authenticate_user
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    if role == 'admin':
        excel_file = EXCEL_FILE
    else:
        excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
    return send_file(excel_file, as_attachment=True)


@app.route("/auto-absent", methods=["POST"])
def auto_absent():
    from utils.user_handler import authenticate_user
    username = session.get('user')
    user = authenticate_user(username, None)
    role = user.get('role', 'teacher') if user else 'teacher'
    batch_list = [b.strip() for b in user.get('batch', '').split(',') if b.strip()] if user else []
    batch_name = request.form.get('batch_name', '').strip()
    if role == 'admin':
        excel_file = EXCEL_FILE
        all_batches = get_batches(excel_file)
    else:
        excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
        all_batches = batch_list
    try:
        batches_to_process = [batch_name] if batch_name else all_batches
        import openpyxl
        from datetime import date
        wb = openpyxl.load_workbook(excel_file)
        sent_count = 0
        total_absentees = []
        today = date.today().strftime("%Y-%m-%d")
        for batch in batches_to_process:
            if batch in wb.sheetnames:
                ws = wb[batch]
                today_col = None
                headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                if today not in headers:
                    ws.cell(row=1, column=ws.max_column + 1, value=today)
                for c in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=c).value == today:
                        today_col = c
                        break
                absentees = []
                for r in range(2, ws.max_row + 1):
                    if ws.cell(row=r, column=today_col).value in (None, ""):
                        ws.cell(row=r, column=today_col, value="Absent")
                        from openpyxl.styles import Font
                        ws.cell(row=r, column=today_col).font = Font(color="FF0000")
                        student_name = ws.cell(row=r, column=2).value or "Student"
                        student_email = ws.cell(row=r, column=3).value or ""
                        if student_email:
                            absentees.append({
                                "name": student_name,
                                "email": student_email,
                                "batch": batch
                            })
                total_absentees.extend(absentees)
        wb.save(excel_file)
        for student in total_absentees:
            if send_absent_email(
                to_email=student["email"],
                student_name=student["name"],
                date_str=str(json.dumps(str(__import__('datetime').date.today()))[1:-1]),
                batch=student["batch"],
                sender_email=SENDER_EMAIL,
                sender_password=SENDER_PASSWORD,
                smtp_server=SMTP_SERVER,
                smtp_port=SMTP_PORT
            ):
                sent_count += 1
        if batch_name:
            flash(f"Auto-marked Absent for batch {batch_name}. Sent {sent_count} absent emails.", "success")
        else:
            flash(f"Auto-marked Absent for all batches. Sent {sent_count} absent emails.", "success")
    except Exception as e:
        flash(f"Error marking absents: {e}", "danger")
    return redirect(url_for("index"))


@app.route("/scan")
def scan_page():
    return render_template("scan.html")


@app.route("/api/scan", methods=["POST"])

def api_scan():
    try:
        payload = request.get_json(force=True, silent=True) or {}
        raw = payload.get("text", "")
        batch = None
        sid = None
        # Try parse JSON first
        try:
            data = json.loads(raw)
            sid = str(data.get("id") or "").strip()
            batch = str(data.get("batch") or "").strip()
        except Exception:
            # Fallback: "BatchName|StudentID"
            if "|" in raw:
                parts = raw.split("|", 1)
                batch = parts[0].strip()
                sid = parts[1].strip()
        if not batch or not sid:
            return jsonify({"ok": False, "msg": "Invalid QR data"}), 400
        # Use correct Excel file for admin/teacher
        from utils.user_handler import authenticate_user
        username = session.get('user')
        user = authenticate_user(username, None)
        role = user.get('role', 'teacher') if user else 'teacher'
        if role == 'admin':
            excel_file = EXCEL_FILE
        else:
            excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
        mark_present(excel_file, batch, sid)
        return jsonify({"ok": True, "msg": f"Marked present: {sid} in {batch}"})
    except Exception as e:
        return jsonify({"ok": False, "msg": f"Error: {e}"}), 500


@app.route("/static/qr_codes/<batch>/<filename>")
def serve_qr(batch, filename):
    return send_from_directory(os.path.join(QR_DIR, batch), filename)


def daily_auto_absent():
    try:
        # Mark absent for admin/global file
        num = mark_absent_all(EXCEL_FILE)
        print(f"[Auto Absent] Marked absent for {num} batch(es) at end of day (admin/global).")
        # Mark absent for all teacher files
        from utils.user_handler import USERS_FILE
        import pandas as pd
        import glob
        df = pd.read_excel(USERS_FILE)
        for _, row in df.iterrows():
            if row.get('role', 'teacher') == 'teacher':
                username = row['username']
                batch = row.get('batch', '')
                excel_file = os.path.join(DATA_DIR, f"{username}_attendance.xlsx")
                if os.path.exists(excel_file) and batch:
                    try:
                        mark_absent_all(excel_file)
                        print(f"[Auto Absent] Marked absent for {username}'s batch: {batch}")
                    except Exception as e:
                        print(f"[Auto Absent] Error for {username}: {e}")
    except Exception as e:
        print(f"[Auto Absent] Error: {e}")

if __name__ == "__main__":
    scheduler = BackgroundScheduler()
    scheduler.add_job(daily_auto_absent, 'cron', hour=23, minute=59)
    scheduler.start()
    app.run(debug=True)
