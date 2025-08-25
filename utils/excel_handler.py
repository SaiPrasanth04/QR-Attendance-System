def delete_student(path: str, batch: str, student_id: str) -> bool:
    """Delete a student by ID from a batch. Returns True if deleted, False if not found."""
    df = read_batch_df(path, batch)
    before = len(df)
    df = df[df['ID'].astype(str) != str(student_id)]
    after = len(df)
    _write_df_back(path, batch, df)
    return after < before
def ensure_email_column_all_batches(path: str):
    """Ensure all batch sheets have an Email column."""
    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "Email" not in headers:
            ws.cell(row=1, column=ws.max_column + 1, value="Email")
    wb.save(path)
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import date

GREEN = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def ensure_workbook(path: str):
    """Create an empty workbook file if it doesn't exist."""
    if not os.path.exists(path):
        wb = Workbook()
        # Optionally rename the default sheet to 'Sheet1' or leave as is
        wb.save(path)


def get_batches(path):
    """Return a list of sheet names (batches) in the Excel file."""
    if not os.path.exists(path):
        # Create empty Excel file with no sheets
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pass
    xl = pd.ExcelFile(path, engine="openpyxl")
    return xl.sheet_names


def _ensure_sheet_headers(wb, sheet_name: str):
    """Ensure a sheet exists with headers ID, Name, Email."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Name")
        ws.cell(row=1, column=3, value="Email")


def read_batch_df(path: str, batch: str) -> pd.DataFrame:
    """Read a batch sheet to DataFrame. If missing, create with headers."""
    wb = load_workbook(path)
    _ensure_sheet_headers(wb, batch)
    wb.save(path)

    df = pd.read_excel(path, sheet_name=batch, engine="openpyxl")
    # Ensure Email column exists
    if "Email" not in df.columns:
        df["Email"] = ""
    # Always sort by Name (case-insensitive) for display
    if "Name" in df.columns:
        try:
            df = df.sort_values(by="Name", key=lambda s: s.str.lower(), ignore_index=True)
        except Exception:
            df = df.sort_values(by="Name", ignore_index=True)
    return df


def _write_df_back(path: str, batch: str, df: pd.DataFrame):
    """Write back one sheet (replace) in the Excel file."""
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=batch, index=False)


def add_student(path: str, batch: str, student_id: str, name: str, email: str) -> bool:
    """
    Add student to batch if ID not present. Stores email as well.
    Returns True if added, False if duplicate ID.
    """

    # Always repair headers before reading
    wb = load_workbook(path)
    if batch not in wb.sheetnames:
        ws = wb.create_sheet(title=batch)
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Name")
        ws.cell(row=1, column=3, value="Email")
        wb.save(path)
    else:
        ws = wb[batch]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        changed = False
        if "ID" not in headers:
            ws.cell(row=1, column=1, value="ID"); changed = True
        if "Name" not in headers:
            ws.cell(row=1, column=2, value="Name"); changed = True
        if "Email" not in headers:
            ws.cell(row=1, column=3, value="Email"); changed = True
        if changed:
            wb.save(path)
    df = read_batch_df(path, batch)
    # Ensure columns exist in DataFrame
    for col in ["ID", "Name", "Email"]:
        if col not in df.columns:
            df[col] = ""

    # Check for duplicate ID and email across all batches
    xl = pd.ExcelFile(path, engine="openpyxl")
    for sheet in xl.sheet_names:
        sdf = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
        # Ensure columns exist in each sheet
        for col in ["ID", "Name", "Email"]:
            if col not in sdf.columns:
                sdf[col] = ""
        if (sdf["ID"].astype(str) == str(student_id)).any():
            return False
        if "Email" in sdf.columns and (sdf["Email"].astype(str).str.lower() == str(email).lower()).any():
            return False

    # Insert row and keep alphabetical by Name
    new_row = {"ID": student_id, "Name": name, "Email": email}
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    try:
        df = df.sort_values(by="Name", key=lambda s: s.str.lower(), ignore_index=True)
    except Exception:
        df = df.sort_values(by="Name", ignore_index=True)

    _write_df_back(path, batch, df)
    return True


def _get_or_create_today_column(ws):
    """Ensure a column exists for today's date, return its index."""
    today = date.today().strftime("%Y-%m-%d")
    header_values = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if today not in header_values:
        ws.cell(row=1, column=ws.max_column + 1, value=today)
    # Find its index
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == today:
            return c


def mark_present(path: str, batch: str, student_id: str):
    """
    Mark Present (green) for today's date for this student ID in the batch sheet.
    """
    wb = load_workbook(path)
    if batch not in wb.sheetnames:
        # Create with headers
        ws = wb.create_sheet(title=batch)
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Name")
    ws = wb[batch]

    date_col = _get_or_create_today_column(ws)

    # Find row with ID
    target_row = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value) == str(student_id):
            target_row = r
            break
    if target_row is None:
        wb.save(path)
        raise ValueError(f"Student ID {student_id} not found in {batch}")

    # Overwrite 'Absent' or empty with 'Present' when scanning
    previous_value = ws.cell(row=target_row, column=date_col).value
    ws.cell(row=target_row, column=date_col, value="Present")
    from openpyxl.styles import Font
    ws.cell(row=target_row, column=date_col).font = Font(color="00B050")  # Green font
    wb.save(path)
    # Send correction email if previously absent
    if previous_value == "Absent":
        try:
            from utils.email_handler import send_absent_email
            from datetime import date
            student_name = ws.cell(row=target_row, column=2).value or "Student"
            student_email = ws.cell(row=target_row, column=3).value or ""
            if student_email:
                import os, sys
                sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
                from app import SENDER_EMAIL, SENDER_PASSWORD, SMTP_SERVER, SMTP_PORT
                send_absent_email(
                    to_email=student_email,
                    student_name=student_name,
                    date_str=str(date.today()),
                    batch=batch,
                    sender_email=SENDER_EMAIL,
                    sender_password=SENDER_PASSWORD,
                    smtp_server=SMTP_SERVER,
                    smtp_port=SMTP_PORT,
                    correction=True
                )
        except Exception as e:
            print(f"Failed to send correction email: {e}")


def mark_absent_all(path: str):
    """
    For each sheet, create today's date column if missing and mark Absent (red)
    for empty cells under today's column.
    Returns (number of sheets processed, list of absent students as dicts).
    """
    wb = load_workbook(path)
    count = 0
    absentees = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Skip empty sheet with no headers
        if ws.max_row < 2 or ws.max_column < 2:
            continue
        date_col = _get_or_create_today_column(ws)
        # Find column indices for Name and Email
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        try:
            name_col = headers.index("Name") + 1
            email_col = headers.index("Email") + 1
        except ValueError:
            name_col = 2
            email_col = 3
        # Mark absents and collect info
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=date_col).value in (None, ""):
                ws.cell(row=r, column=date_col, value="Absent")
                from openpyxl.styles import Font
                ws.cell(row=r, column=date_col).font = Font(color="FF0000")  # Red font
                student_name = ws.cell(row=r, column=name_col).value or "Student"
                student_email = ws.cell(row=r, column=email_col).value or ""
                if student_email:
                    absentees.append({
                        "name": student_name,
                        "email": student_email,
                        "batch": sheet
                    })
        count += 1
    wb.save(path)
    return count, absentees


def download_path(base_dir: str, filename: str) -> str:
    """Return full path for download file."""
    return os.path.join(base_dir, filename)
